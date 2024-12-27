from openpyxl import load_workbook

class SaveExel():
    def __init__(self):
        pass

    def save(self, _file, _sheet, _name, quantity):
        try :
            wb = load_workbook(_file)
            ws = wb[_sheet]
            column = ws['A']
            for cell in column:
                if cell.value is not None and cell.value == _name:
                    _row = ws[cell.row]
                    for _cell in _row[-2::]:
                        _cell.value = quantity
                        wb.save(_file)
                        return True
        except Exception:
            return False

if __name__ == '__main__':
    se = SaveExel()
    if se.save("docx/docxxx1.xlsx", 'MMQ  HOSPITAL ', 'ADHESIVO', 445):
        print("Se logro hacer el cambio en el exel")
    else:
        print("Se logro hacer el cambio en el exel")
